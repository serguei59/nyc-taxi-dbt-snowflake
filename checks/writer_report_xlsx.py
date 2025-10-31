# verifications/writer_report_xlsx.py

from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path

def save_ingestion_report_xlsx(stats: dict):
    """
    Save post-ingestion data quality checks into an Excel file.
    - Summary sheet accumulates all runs
    - A separate sheet is created for each run
    """
    report_dir = Path(__file__).parent
    report_dir.mkdir(parents=True, exist_ok=True)
    report_file = report_dir / "post_ingestion_report.xlsx"

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet_name = f"Run_{datetime.now().strftime('%Y%m%d_%H%M')}"

    # Define columns matching the DQ checks
    headers = [
        "timestamp",
        "total_rows",
        "duplicate_groups",
        "buffer_rows",
        "min_distance",
        "max_distance",
        "avg_distance"
    ]

    # Build row from stats
    row = [
        timestamp,
        stats.get("TOTAL_ROWS", "N/A"),
        stats.get("DUPLICATE_GROUPS", "N/A"),
        stats.get("BUFFER_ROWS", "N/A"),
        stats.get("MIN_DISTANCE", "N/A"),
        stats.get("MAX_DISTANCE", "N/A"),
        stats.get("AVG_DISTANCE", "N/A"),
    ]

    # Load or create workbook
    if report_file.exists():
        wb = load_workbook(report_file)
    else:
        wb = Workbook()

    # --- 1️⃣ Summary sheet
    if "Summary" not in wb.sheetnames:
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary.append(headers)
    else:
        ws_summary = wb["Summary"]
    ws_summary.append(row)

    # --- 2️⃣ Detailed sheet for this run
    ws_run = wb.create_sheet(title=sheet_name)
    ws_run.append(headers)
    ws_run.append(row)

    # Save workbook
    wb.save(report_file)

    print(f"📊 Ingestion report saved: {report_file}")
    print(f"🗂️  New sheet created: {sheet_name}")

# Example usage in merge_dynamic.py:
# save_ingestion_report_xlsx(results)
