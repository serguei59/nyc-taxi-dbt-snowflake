from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path

def write_merge_report(file_name, rows_merged, status, report_dir="reports"):
    report_path = Path(report_dir)
    report_path.mkdir(parents=True, exist_ok=True)
    report_file = report_path / "merge_report.xlsx"

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    headers = ["timestamp", "file_name", "rows_merged", "status"]
    row = [timestamp, file_name, rows_merged, status]

    wb = load_workbook(report_file) if report_file.exists() else Workbook()
    ws = wb.active
    ws.title = "Summary"
    if ws.max_row == 1 and ws["A1"].value != "timestamp":
        ws.append(headers)
    ws.append(row)
    wb.save(report_file)
    print(f"✅ Report updated: {report_file}")
